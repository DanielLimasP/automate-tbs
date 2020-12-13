import openpyxl
from pathlib import Path  
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

print(Path.cwd())
wb = openpyxl.load_workbook(Path.cwd() / 'spreadsheets/example.xlsx')
type(wb) # --> print(type(wb))
# <class 'openpyxl.workbook.workbook.Workbook'>
wb.sheetnames
# ['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3']
type(sheet)
# <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title
# 'Sheet3'
another_sheet = wb.active
another_sheet
# <Worksheet "Sheet1">

sheet = wb['Sheet1']
sheet.title
# 'Sheet1'
sheet['A1']
# <Cell 'Sheet1'.A1>
sheet['A1'].value
# datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1']
c.value
# 'Apples'
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
# 'Row 1, Column 2 is Apples'
print('Cell {} is {}'.format(c.coordinate, c.value))
# 'Cell B1 is Apples'
sheet['C1']
# <Cell 'Sheet1'.C1>
sheet['C1'].value
# 73

sheet.cell(row=1, column=2)
# <Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value
# 'Apples'
for i in range(1, 8, 2): # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)
    
sheet.max_row # Get the highest row number.
sheet.max_column # Get the highest column number.

# Convert from column letters to numbers and viceversa
get_column_letter(1) # Translate column 1 to a letter.
# 'A'
get_column_letter(2)
# 'B'
get_column_letter(27)
# 'AA'
get_column_letter(900)
# 'AHP'
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
# 'C'
column_index_from_string('A') # Get A's number.
# 1
column_index_from_string('AA')
# 27

# Better loop with a tuple of the items in a range of columns
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
#((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell
#'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>,
#<Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# Alternatively, we can use the rows and columns attributes of the sheet to get specific values 
list(sheet.columns)[1] # Get second column's cells.
# (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.
# B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)
for cellObj in list(sheet.columns)[1]:
        print(cellObj.value)

# Creating new spreadsheets
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx') # Save the workbook.

# Creating new sheets in a workbook
wb.create_sheet() # Add a new sheet.
# <Worksheet "Sheet1">
wb.sheetnames
# ['Sheet', 'Sheet1']
# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')
# <Worksheet "First Sheet">
wb.sheetnames
# ['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')
# <Worksheet "Middle Sheet">
wb.sheetnames

# you can use python del to get rid of sheets
del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames

# Like this we can add new values to our spreadsheet
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!' # Edit the cell's value.
sheet['A1'].value

# Changing the font...
wb = openpyxl.Workbook()
sheet = wb.active

font_obj_1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = font_obj_1
sheet['A1'] = 'Bold Times New Roman'

font_obj_2 = Font(size=24, italic=True)
sheet['B3'].font = font_obj_2
sheet['B3'] = '24 pt Italic'

wb.save('spreadsheets/styles.xlsx')

# Additionally, we can write formulas to cells as if they were normal values
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
wb.save('spreadsheet/write_formula.xlsx')

# We can even set width to cols and height to rows
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
# Set the height and width:
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('spreadsheets/dimensions.xlsx')

# to merge cells...
sheet = wb.active
sheet.merge_cells('A1:D3') # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5') # Merge these two cells.
sheet['C5'] = 'Two merged cells.'
wb.save('spreadsheets/merged.xlsx')
# and to unmerge them...
sheet = wb.active
sheet.unmerge_cells('A1:D3') # Split these cells up.
sheet.unmerge_cells('C5:D5')
wb.save('spreadsheets/merged.xlsx')

# You can "freeze" panes
sheet = wb.active
sheet.freeze_panes = 'A2' # Freeze the rows above A2.
wb.save('spreadsheets/freezeExample.xlsx')

def create_chart():
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, 11):
        sheet['A' + str(i)] = i
    
reference_object = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
series_object = openpyxl.chart.Series(reference_object, title='First Series')
chart_object = openpyxl.chart.BarChart()
chart_object.title = 'My Chart'
chart_object.append(series_object)

sheet.add_chart(chart_object, 'C5')
wb.save('spreadsheets/sample_chart.xlsx')