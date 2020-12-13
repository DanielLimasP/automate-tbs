#! python3
# update_products_prices.py - Corrects costs in produce sales spreadsheet

import openpyxl
wb = openpyxl.load_workbook('spreadsheets/products_sales.xlsx')

sheet = wb.active

# This is our update data structure
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

for row_num in range(2, sheet.max_row):
    product_name = sheet.cell(row=row_num, column=1).value
    if product_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[product_name]

wb.save('spreadsheets/updated_product_sales.xlsx')

